# -*- coding:utf-8 -*-
#!/usr/bin/python3
# @Time: 2021/1/26 5:16 下午
# @Author: zhifeng.qin
# @File: 读取 Excel 数据.py
# @Software: PyCharm

'''
    读取 excel 文件数据
'''

import openpyxl
class ReadExcel:

    def get_excel(self):

        # wb 打开工作目录
        wb = openpyxl.load_workbook('./login_data.xlsx')

        # ws sheet 页，读取对应表单, 登录数据为对应 sheet 表单的名称
        ws = wb['登录数据']

        # 构造数据  [[],[],[],[]]
        all_case = []

        # 获取对应表单所有数据（通过生成器）
        select_all_case = ws.iter_rows(min_row = 1, max_row = ws.max_row, min_col = 1, max_col = ws.max_column)

        # 循环遍历所有 case，然后拿到表格中一行中数据（即一个列表），把列表添加到新的列表中，最终返回结果是列表套列表
        for rows in select_all_case:
            cell_value_list = [cell.value for cell in rows]     # 得到表单中一行的数据
            all_case.append(cell_value_list)
        return all_case

if __name__ == '__main__':
    print(ReadExcel().get_excel())






